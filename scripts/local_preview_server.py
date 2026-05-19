from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
import base64
import csv
import gzip
import io
import json
import os
from pathlib import Path
import re
import threading
import time
from typing import Any
import urllib.error
import urllib.request
from urllib.parse import parse_qs, unquote, urlencode, urlparse
import uuid
import zipfile

import snowflake.connector

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


ROOT = Path(__file__).resolve().parents[1]
PORT = int(os.environ.get("PORT", "4200"))
LEGACY_REDIRECT_PORT = int(os.environ.get("LEGACY_REDIRECT_PORT", "4173") or "0")


def load_local_env() -> None:
    for name in (".env.local", ".env"):
        path = ROOT / name
        if not path.exists():
            continue
        for raw_line in path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            if key in os.environ:
                continue
            os.environ[key] = value.strip().strip("\"'")


load_local_env()

AUTH_PROFILE_CACHE_TTL_SECONDS = int(os.environ.get("AUTH_PROFILE_CACHE_TTL_SECONDS", "60") or "60")
_auth_profile_cache: dict[str, tuple[float, dict[str, Any]]] = {}
_auth_profile_cache_lock = threading.Lock()


def truthy_env(name: str) -> bool:
    return str(os.environ.get(name, "")).strip().lower() in {"1", "true", "yes", "on"}


def first_env(*names: str) -> str:
    for name in names:
        value = os.environ.get(name)
        if value and value.strip():
            return value.strip()
    return ""


def with_leading_slash(value: str, fallback: str) -> str:
    raw = (value or fallback).strip()
    return raw if raw.startswith("/") else f"/{raw}"


def auth_required() -> bool:
    return False


def auth_authority() -> str:
    return first_env("OIDC_AUTHORITY", "OIDC_ISSUER", "AUTH_AUTHORITY").rstrip("/")


def auth_authorize_url() -> str:
    explicit = first_env("OIDC_AUTHORIZE_URL", "AUTH_AUTHORIZE_URL")
    if explicit:
        return explicit
    authority = auth_authority()
    return f"{authority}/authorize" if authority else ""


def auth_client_id() -> str:
    return first_env("OIDC_CLIENT_ID", "AUTH_CLIENT_ID")


def auth_client_secret() -> str:
    return first_env("OIDC_CLIENT_SECRET", "AUTH_CLIENT_SECRET")


def auth_redirect_uri() -> str:
    return first_env("OIDC_LOCAL_REDIRECT_URI") or first_env("OIDC_REDIRECT_URI", "AUTH_REDIRECT_URI") or f"http://localhost:{PORT}"


def local_request_origin(headers) -> str:
    host = headers.get("host") or headers.get("Host") or f"127.0.0.1:{PORT}"
    return f"http://{host}"


def redirect_uri_for(headers=None) -> str:
    configured = auth_redirect_uri()
    origin = local_request_origin(headers) if headers is not None else f"http://127.0.0.1:{PORT}"
    host = (headers.get("host") or headers.get("Host") or "") if headers is not None else ""
    local_host = host.startswith("localhost:") or host.startswith("127.0.0.1:")
    if local_host:
        return origin
    if not configured or "YOUR_APP.vercel.app" in configured:
        return origin
    return configured


def auth_token_path() -> str:
    return with_leading_slash(first_env("OIDC_TOKEN_PATH", "AUTH_TOKEN_PATH"), "/token")


def auth_profile_path() -> str:
    return with_leading_slash(first_env("OIDC_PROFILE_PATH", "AUTH_PROFILE_PATH"), "/profile")


def auth_logout_url() -> str:
    authority = auth_authority()
    return first_env("OIDC_LOGOUT_URL", "AUTH_LOGOUT_URL") or (f"{authority}/logout" if authority else "")


def public_auth_config(headers=None) -> dict[str, Any]:
    return {
        "enabled": False,
        "required": False,
        "configured": False,
        "oidcIssuer": "",
        "authorizeUrl": "",
        "clientId": "",
        "redirectUri": redirect_uri_for(headers),
        "logoutUrl": "",
    }


def exchange_auth_code(code: str, redirect_uri: str | None = None) -> dict[str, Any]:
    authority = auth_authority()
    client_id = auth_client_id()
    client_secret = auth_client_secret()
    if not authority or not client_id or not client_secret:
        raise PermissionError("OIDC authority, client id, and client secret are required for code exchange.")
    if not code or not str(code).strip():
        raise PermissionError("Authorization code is required.")

    body = urlencode(
        {
            "grant_type": "authorization_code",
            "code": code,
            "client_id": client_id,
            "client_secret": client_secret,
            "redirect_uri": redirect_uri or redirect_uri_for(),
        }
    ).encode("utf-8")
    request = urllib.request.Request(
        f"{authority}{auth_token_path()}",
        data=body,
        headers={"content-type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8") or "{}")
    except urllib.error.HTTPError as exc:
        try:
            error_payload = json.loads(exc.read().decode("utf-8") or "{}")
        except Exception:
            error_payload = {}
        message = error_payload.get("error_description") or error_payload.get("error") or f"OIDC code exchange failed with {exc.code}."
        raise PermissionError(message) from exc

    if not payload.get("access_token"):
        raise PermissionError("OIDC code exchange did not return an access token.")
    return {
        "access_token": payload.get("access_token"),
        "id_token": payload.get("id_token"),
        "token_type": payload.get("token_type"),
        "expires_in": payload.get("expires_in"),
    }


def validate_token_via_profile(token: str) -> dict[str, Any] | None:
    if not token:
        return None
    bypass = first_env("AUTH_BYPASS_TOKEN")
    if bypass and token == bypass:
        return {"sub": "bypass", "name": "Token Bypass"}

    now = time.time()
    with _auth_profile_cache_lock:
        cached = _auth_profile_cache.get(token)
        if cached and now - cached[0] < AUTH_PROFILE_CACHE_TTL_SECONDS:
            return cached[1]
        _auth_profile_cache.pop(token, None)

    authority = auth_authority()
    if not authority:
        return None

    profile_url = f"{authority}{auth_profile_path()}?{urlencode({'access_token': token})}"
    try:
        with urllib.request.urlopen(profile_url, timeout=20) as response:
            claims = json.loads(response.read().decode("utf-8") or "{}")
    except Exception:
        return None

    if isinstance(claims, dict):
        with _auth_profile_cache_lock:
            _auth_profile_cache[token] = (now, claims)
        return claims
    return None


def bearer_token(headers) -> str:
    value = headers.get("authorization") or headers.get("Authorization") or ""
    return value[7:].strip() if value.lower().startswith("bearer ") else ""


def is_public_api_path(path: str) -> bool:
    return path.lower() in {"/api/auth/config", "/api/auth/exchange-code", "/api/health"}


def is_authorized(headers) -> bool:
    return True

REPORT_DB = os.environ.get("SNOWFLAKE_DATABASE") or os.environ.get("SNOWFLAKE_DB") or "FOODBUY_MASALA_PROD"
REPORT_SCHEMA = os.environ.get("SNOWFLAKE_SCHEMA") or "COMPLIANCE_LAB"
SILVER_SCHEMA = os.environ.get("SNOWFLAKE_SILVER_SCHEMA") or "MASALA_SILVER_COMPLIANCE_LAB"
OPENSTOCK_KEY = "DISTCODE MOG DIN"
OPENSTOCK_TABLE = f"{REPORT_DB}.{REPORT_SCHEMA}.OPENSTOCKREPORT"
CHANGE_BATCH_TABLE = f"{REPORT_DB}.{REPORT_SCHEMA}.OPENSTOCK_CHANGE_BATCH"
CHANGE_LOG_TABLE = f"{REPORT_DB}.{REPORT_SCHEMA}.OPENSTOCK_CHANGE_LOG"
FEEDBACK_HUB_TABLE = f"{REPORT_DB}.{REPORT_SCHEMA}.FEEDBACK_HUB"
UA_AUDIT_LOG_TABLE = f"{REPORT_DB}.{REPORT_SCHEMA}.UNLOCKED_ACCOUNTS_AUDIT_LOG"
UA_HISTORY_TABLE = f"{REPORT_DB}.{REPORT_SCHEMA}.UNLOCKED_ACCOUNTS_HISTORY"
DISTRIBUTOR_HIERARCHY_TABLE = "FOODBUY_IRON_GURU_PROD.READER.DISTRIBUTORHIERARCHY"
SUPER_USERS = {"jordaa14", "phillg02", "gilbem02", "sullik09"}

OPENSTOCK_EDITABLE_COLS = [
    "New Item?",
    "In Stock (Y/N?)",
    "ETA",
    "PO #",
    "Current DC Comment",
    "Current SCS Comment",
    "Required DC Update",
    "Pending Management Comments",
]
OPENSTOCK_UPLOAD_EDITABLE_COLS = [*OPENSTOCK_EDITABLE_COLS, "SCS"]
OPENSTOCK_PREVIOUS_COLUMN_MAP = {
    "Previous DC Comment": "Current DC Comment",
    "Previous SCS Comment": "Current SCS Comment",
    "Previous DC Update": "Required DC Update",
}
OPENSTOCK_CARRY_FORWARD_COLS = [
    "In Stock (Y/N?)",
    "ETA",
    "PO #",
    "Current DC Comment",
    "Current SCS Comment",
    "Required DC Update",
    "Pending Management Comments",
]

DC_MATRIX_LOCKED_COLS = [
    "SC_PARENT_NAME",
    "DISTRIBUTOR_TYPE",
    "DISTRIBUTOR_COUNTRY",
    "SUPPLY_CHAIN_NAME",
    "SUPPLY_CHAIN_PARENT_CODE",
    "SUPPLY_CHAIN_CODE",
]
DC_MATRIX_EDITABLE_COLS = [
    "MOG_TYPE",
    "ITRADE_NAME",
    "REACTORNET_NAME",
    "SHORT_NAME",
    "ITRADE_PARENT_CODE",
    "ITRADE_CODE",
    "COMPASS",
    "HMS_HOST",
    "CSM",
    "AIMBRIDGE",
    "HEALTHTRUST",
    "COMPASS_CONTROLLED_DC",
    "CONVERSION_DC",
    "REMEDY_CODE",
    "PHASE_OUT_NAME",
    "FUTURECARE",
    "ALTERNATE_NAME",
    "NAME_CONV_DIST_DC_NAME",
    "NAME_CONV_DIST_PARENT",
    "NAME_CONV_SC_DC_NAME",
]

UNLOCKED_TEMPLATE_COLS = [
    "BUSINESS",
    "CUSTOMER",
    "DC_NAME",
    "DISTRIBUTOR_CODE",
    "SECTOR_ATTRIBUTE",
    "UNIT_NUMBER",
    "DSTCODEUNIT",
    "CUS_CODE",
    "DIST_CUSTOMER_NAME",
    "DCN",
    "DSTCODEDCN",
    "DCN_CODE",
    "DATE_UNLOCKED",
    "UNLOCK_DOTCOM",
    "UNLOCK_MYORDERS",
    "UNLOCK_NO_CAT_MYORDERS",
    "ACCOUNT_TYPE",
    "REQUESTOR_NAME",
]


def fqn(object_name: str, schema: str = REPORT_SCHEMA) -> str:
    return f"{REPORT_DB}.{schema}.{object_name}"


def source(
    object_name: str,
    schema: str = REPORT_SCHEMA,
    key_columns: list[str] | None = None,
    snapshot_columns: list[str] | None = None,
    candidates: list[str] | None = None,
    label: str | None = None,
) -> dict[str, Any]:
    object_names = [fqn(object_name, schema)]
    if candidates:
        object_names.extend(candidates)
    return {
        "name": object_name,
        "label": label or object_name,
        "objectName": object_names[0],
        "objectNames": object_names,
        "keyColumns": key_columns or [],
        "snapshotColumns": snapshot_columns or [],
    }


def hub(
    label: str,
    description: str,
    sources: list[dict[str, Any]],
    columns: list[str],
    filter_columns: list[str] | None = None,
    editable_columns: list[str] | None = None,
    search_columns: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "label": label,
        "description": description,
        "sources": sources,
        "columns": columns,
        "editableColumns": editable_columns or [],
        "filterColumns": filter_columns or [],
        "searchColumns": search_columns or [],
    }


HUBS: dict[str, dict[str, Any]] = {
    "open-stock": {
        "label": "Open Stock",
        "description": "Editable Open Stock worklist with saved change history.",
        "sources": [
            source("OPENSTOCKREPORT", key_columns=[OPENSTOCK_KEY], snapshot_columns=["INSERT_DATE"]),
            source("OPENSTOCKREPORT_ALLMOG", key_columns=[OPENSTOCK_KEY], snapshot_columns=["INSERT_DATE"]),
            source("OPENSTOCKREPORT_OS", key_columns=[OPENSTOCK_KEY], snapshot_columns=["INSERT_DATE"]),
            source("OPENSTOCKREPORT_SYSCO", key_columns=[OPENSTOCK_KEY], snapshot_columns=["INSERT_DATE"]),
        ],
        "columns": [
            "DISTRIBUTOR NAME",
            "DISTRIBUTOR ID",
            OPENSTOCK_KEY,
            "MOG NAME",
            "MOG FLAG DESC",
            "MANUFACTURER NAME",
            "MIN",
            "BRAND",
            "DESCRIPTION",
            "PACK SIZE",
            "CREATED DATE",
            "DIN",
            "New Item?",
            "In Stock (Y/N?)",
            "ETA",
            "+2 Weeks",
            "PO #",
            "Previous DC Comment",
            "Required DC Update",
            "Current DC Comment",
            "Previous SCS Comment",
            "Current SCS Comment",
            "Pending Management Comments",
            "SCS",
        ],
        "editableColumns": OPENSTOCK_EDITABLE_COLS,
        "filterColumns": ["DISTRIBUTOR NAME", "SCS", "In Stock (Y/N?)", "New Item?", "+2 Weeks"],
        "searchColumns": [
            "DISTRIBUTOR NAME",
            "SCS",
            "MANUFACTURER NAME",
            "BRAND",
            "DESCRIPTION",
            "DIN",
            OPENSTOCK_KEY,
            "Current DC Comment",
            "Current SCS Comment",
            "Required DC Update",
            "Pending Management Comments",
            "PO #",
        ],
    },
    "dc-matrix": hub(
        "DC Matrix",
        "DC Matrix records and supply-chain mapping.",
        [source("V_DC_MATRIX", key_columns=["SUPPLY_CHAIN_CODE"], label="DC Matrix")],
        [*DC_MATRIX_LOCKED_COLS, *DC_MATRIX_EDITABLE_COLS],
        ["DISTRIBUTOR_TYPE", "SUPPLY_CHAIN_NAME", "SUPPLY_CHAIN_CODE"],
        editable_columns=DC_MATRIX_EDITABLE_COLS,
        search_columns=["SC_PARENT_NAME", "SUPPLY_CHAIN_NAME", "SUPPLY_CHAIN_CODE", "ITRADE_NAME", "SHORT_NAME"],
    ),
    "conversions": hub(
        "Conversions",
        "Conversion read models, action files, and manual workflow tables.",
        [
            source("V_WORKING_MASTER_TOOL", key_columns=["PrimaryKey", "PRIMARYKEY"], label="Working Master Tool"),
            source("V_ACTION_FILE_TOOL", key_columns=["PrimaryKey", "PRIMARYKEY"], label="Action File Tool"),
            source("SOURCING_CONVERSION_MASTER_TBL", key_columns=["PrimaryKey", "PRIMARYKEY"], label="Sourcing Conversion Master"),
            source("CONVERSION_ANALYSIS_SRFS_MASTER_TBL", key_columns=["PrimaryKey", "PRIMARYKEY", "ROW #"], label="Conversion Master"),
            source("DC_COMMUNICATIONS_MANUAL_TBL", key_columns=["PrimaryKey", "PRIMARYKEY", "SEQUENCE"], label="DC Communications Manual"),
            source("DC_COMMUNICATION_TOOL", SILVER_SCHEMA, key_columns=["SEQUENCE"], label="DC Communications View"),
        ],
        ["PrimaryKey", "ConversionMonth", "DISTRIBUTOR NAME", "ACTION", "DATE EXECUTED (ACTUAL DATE)", "COMPLETION STATUS", "COMPLETION COMMENTS", "Analyst"],
        ["ConversionMonth", "DISTRIBUTOR NAME", "ACTION", "COMPLETION STATUS", "Analyst"],
        editable_columns=[
            "Reversed Sectors",
            "Requested By",
            "Reversal Reason",
            "Date Implemented",
            "DATE EXECUTED (ACTUAL DATE)",
            "NEW ITEM ATTRIBUTES",
            "NEW MOG",
            "CONVERSION ANALYSIS COMMENTS",
            "COMPLETION STATUS",
            "COMPLETION COMMENTS",
            "ROW #",
            "Analyst",
        ],
        search_columns=["PrimaryKey", "DISTRIBUTOR NAME", "ACTION", "CONVERSION ANALYSIS NAME", "ITEM DESCRIPTION", "Analyst"],
    ),
    "unlocked-accounts": hub(
        "Unlocked Accounts",
        "Current unlocked and locked account state.",
        [
            source("UNLOCKED_ACCOUNTS", key_columns=["ACCOUNT_RECORD_ID", "DSTCODEDCN"], candidates=[fqn("UNLOCKED_ACCOUNTS", SILVER_SCHEMA)], label="Unlocked Accounts"),
            source("LOCKED_INACTIVE_ACCOUNTS", key_columns=["ACCOUNT_RECORD_ID", "DSTCODEDCN"], candidates=[fqn("LOCKED_INACTIVE_ACCOUNTS", SILVER_SCHEMA)], label="Locked Accounts"),
        ],
        [*UNLOCKED_TEMPLATE_COLS, "DATE_LOCKED", "LOCK_REASON", "LAST_TRANSACTION_DATE", "LAST_TRANSACTION_REFRESHED_AT"],
        ["DCN", "UNIT_NUMBER", "DC_NAME", "SECTOR_ATTRIBUTE"],
        editable_columns=["SECTOR_ATTRIBUTE", "CUS_CODE", "DIST_CUSTOMER_NAME", "DCN_CODE", "DATE_UNLOCKED", "UNLOCK_DOTCOM", "UNLOCK_MYORDERS", "UNLOCK_NO_CAT_MYORDERS", "ACCOUNT_TYPE", "REQUESTOR_NAME"],
        search_columns=["BUSINESS", "CUSTOMER", "DC_NAME", "DISTRIBUTOR_CODE", "UNIT_NUMBER", "DSTCODEDCN", "REQUESTOR_NAME"],
    ),
    "slow-dead": hub(
        "Slow and Dead",
        "Slow and dead inventory view with sector and category analysis.",
        [source("V_SLOWDEAD_ALL", key_columns=["DISTCODEDIN", "DIN", "MIN"], label="All S&D")],
        ["Sector", "Category", "NOTICE", "QOH", "True Extended Value", "Intentional?"],
        ["Sector", "Category", "NOTICE"],
        search_columns=["Sector", "Category", "NOTICE", "DESCRIPTION", "BRAND", "MIN"],
    ),
    "itrade": hub(
        "iTrade",
        "iTrade reference views loaded as separate source tabs.",
        [
            source("V_ITRADE_ACCOUNT_LIST", label="Account List"),
            source("V_AUTOSHIPMENT_ITRADE_TOOL", label="Autoshipments iTrade Tool"),
            source("V_ITRADE_CONVERSION_BAR_UNITS", label="Conversion BAR Units"),
            source("V_ITRADE_SECTORS_AT_DC", candidates=[fqn("V_SECTORS_AT_DC")], label="Sectors at DC"),
        ],
        ["ACCOUNT", "DISTRIBUTOR", "SECTOR", "STATUS"],
        ["SECTOR", "DISTRIBUTOR", "STATUS"],
    ),
    "off-mog": hub("Off MOG", "Off MOG reference view.", [source("V_OFF_MOG", label="Off MOG")], ["DISTRIBUTOR", "MOG", "DIN", "BRAND", "DESCRIPTION"], ["DISTRIBUTOR", "MOG"]),
    "prop-list": hub("Prop List", "Monthly proprietary list view.", [source("V_PROPRIETARY_LIST_MONTHLY", key_columns=["DIN", "MIN"], label="Prop List")], ["SECTOR", "NOTICE", "CATEGORY", "DIN", "MIN", "BRAND", "DESCRIPTION"], ["SECTOR", "NOTICE", "CATEGORY"]),
    "substitutions": hub("Substitutions", "Substitutions view with global search and export.", [source("V_SUBSTITUTIONS", label="Substitutions")], ["DISTRIBUTOR", "DIN", "BRAND", "DESCRIPTION", "SUBSTITUTE"], ["DISTRIBUTOR", "CATEGORY"]),
    "autoshipments": hub("Autoshipments", "Autoshipments workflow view.", [source("V_AUTO_SHIPMENTS", label="Autoshipments")], ["SUBMISSION MONTH", "SUBMISSION DAY", "SUBMISSION YEAR", "ISSUES FOUND", "ACCOUNT", "DISTRIBUTOR"], ["SUBMISSION MONTH", "SUBMISSION DAY", "SUBMISSION YEAR", "ISSUES FOUND"]),
}

_connection = None
_connection_lock = threading.Lock()
_snowflake_io_lock = threading.Lock()
_column_cache: dict[str, dict[str, str]] = {}
_payload_cache: dict[str, tuple[float, dict[str, Any]]] = {}
_payload_cache_lock = threading.Lock()
PAYLOAD_CACHE_TTL_SECONDS = int(os.environ.get("LOCAL_PREVIEW_CACHE_TTL_SECONDS", "180"))
PAYLOAD_CACHE_LIMIT = int(os.environ.get("LOCAL_PREVIEW_CACHE_LIMIT", "10"))


def quote_ident(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def insert_date_expr(column_sql: str) -> str:
    return f"SUBSTR(REGEXP_REPLACE(TO_VARCHAR({column_sql}), '[^0-9]', ''), 1, 8)"


def normalize_snapshot(value: Any, fallback: str = "current") -> str:
    text = str(value or "").strip()
    digits = re.sub(r"\D", "", text)
    return digits[:8] if len(digits) >= 8 else text or fallback


def snowflake_connect():
    global _connection
    with _connection_lock:
        try:
            if _connection is not None and not _connection.is_closed():
                return _connection
        except Exception:
            _connection = None

        account = os.environ.get("SNOWFLAKE_ACCOUNT")
        user = os.environ.get("SNOWFLAKE_USERNAME")
        warehouse = os.environ.get("SNOWFLAKE_WAREHOUSE")
        authenticator = os.environ.get("SNOWFLAKE_AUTHENTICATOR") or "externalbrowser"
        missing = [name for name, value in {
            "SNOWFLAKE_ACCOUNT": account,
            "SNOWFLAKE_USERNAME": user,
            "SNOWFLAKE_WAREHOUSE": warehouse,
        }.items() if not value]
        if missing:
            raise RuntimeError("Snowflake connection is not configured. Missing: " + ", ".join(missing))

        kwargs = {
            "account": account,
            "user": user,
            "warehouse": warehouse,
            "database": REPORT_DB,
            "schema": REPORT_SCHEMA,
            "authenticator": authenticator,
        }
        if os.environ.get("SNOWFLAKE_ROLE"):
            kwargs["role"] = os.environ["SNOWFLAKE_ROLE"]
        if os.environ.get("SNOWFLAKE_PASSWORD") and authenticator.lower() not in {"externalbrowser", "oauth"}:
            kwargs["password"] = os.environ["SNOWFLAKE_PASSWORD"]
        if authenticator.lower() == "oauth":
            kwargs["token"] = os.environ.get("SNOWFLAKE_OAUTH_TOKEN") or os.environ.get("SNOWFLAKE_TOKEN")
        if os.environ.get("SNOWFLAKE_CLIENT_STORE_TEMPORARY_CREDENTIAL", "").lower() == "true":
            kwargs["client_store_temporary_credential"] = True

        _connection = snowflake.connector.connect(**kwargs)
        return _connection


def snowflake_query(sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    with _snowflake_io_lock:
        conn = snowflake_connect()
        with conn.cursor(snowflake.connector.DictCursor) as cur:
            cur.execute(sql, params)
            if cur.description is None:
                return []
            return list(cur.fetchall())


def snowflake_execute(sql: str, params: tuple[Any, ...] = ()) -> int:
    with _snowflake_io_lock:
        conn = snowflake_connect()
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return int(cur.rowcount or 0)


def object_exists(object_name: str) -> bool:
    parts = object_name.split(".")
    if len(parts) != 3:
        return False
    rows = snowflake_query(
        f"""
        select 1
        from {parts[0]}.information_schema.columns
        where table_schema = %s
          and table_name = %s
        limit 1
        """,
        (parts[1].upper(), parts[2].upper()),
    )
    return bool(rows)


def resolve_source_object(source_config: dict[str, Any]) -> str:
    cached = source_config.get("resolvedObjectName")
    if cached:
        return str(cached)
    for object_name in source_config.get("objectNames") or [source_config["objectName"]]:
        if object_exists(str(object_name)):
            source_config["resolvedObjectName"] = str(object_name)
            source_config["objectName"] = str(object_name)
            return str(object_name)
    return str(source_config["objectName"])


def source_columns(source_config: dict[str, Any]) -> dict[str, str]:
    object_name = resolve_source_object(source_config)
    if object_name in _column_cache:
        return _column_cache[object_name]
    parts = object_name.split(".")
    if len(parts) != 3:
        return {}
    rows = snowflake_query(
        f"""
        select column_name
        from {parts[0]}.information_schema.columns
        where table_schema = %s
          and table_name = %s
        order by ordinal_position
        """,
        (parts[1].upper(), parts[2].upper()),
    )
    columns = {str(row["COLUMN_NAME"]).upper(): str(row["COLUMN_NAME"]) for row in rows}
    _column_cache[object_name] = columns
    return columns


def actual_column(columns: dict[str, str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        found = columns.get(candidate.upper())
        if found:
            return found
    return None


def first_present(row: dict[str, Any], candidates: list[str]) -> Any:
    by_upper = {key.upper(): key for key in row}
    for candidate in candidates:
        actual = by_upper.get(candidate.upper(), candidate)
        value = row.get(actual)
        if value is not None and str(value).strip() != "":
            return value
    return None


def row_key_for(row: dict[str, Any], source_config: dict[str, Any]) -> str:
    value = first_present(row, source_config.get("keyColumns") or [])
    if value is not None:
        return re.sub(r"\s+", " ", str(value).strip())
    raw = json.dumps(json_safe(row), sort_keys=True).encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii")[:32]


def filter_values(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    if isinstance(value, str) and value.strip() and value != "All":
        return [value]
    return []


def normalize_filters(raw: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(raw)
    aliases = {
        "distributor": "DISTRIBUTOR NAME",
        "scs": "SCS",
        "inStock": "In Stock (Y/N?)",
        "newItem": "New Item?",
    }
    for source_key, target_col in aliases.items():
        values = filter_values(raw.get(source_key))
        if values:
            normalized[target_col] = values
    return normalized


def parse_page_size(query: dict[str, list[str]]) -> int | None:
    raw_value = str(query.get("pageSize", ["all"])[0] or "all").strip().lower()
    if raw_value in {"all", "none", "full", "0", "-1"}:
        return None
    try:
        return max(1, int(raw_value))
    except ValueError:
        return None


def build_rows_for_source(config: dict[str, Any], source_config: dict[str, Any], query: dict[str, list[str]]) -> list[dict[str, Any]]:
    columns = source_columns(source_config)
    object_name = resolve_source_object(source_config)
    filters = normalize_filters(json.loads(query.get("filters", ["{}"])[0] or "{}"))
    run_date = query.get("runDate", [""])[0]
    search = query.get("search", [""])[0].strip().lower()
    page_size = parse_page_size(query)
    page = max(1, int(query.get("page", ["1"])[0] or "1"))
    clauses: list[str] = []
    params: list[Any] = []

    snapshot_col = actual_column(columns, source_config.get("snapshotColumns") or [])
    if run_date and snapshot_col:
        clauses.append(f"{insert_date_expr(quote_ident(snapshot_col))} = %s")
        params.append(normalize_snapshot(run_date))

    allowed_filters = {col.upper() for col in config["filterColumns"]}
    for column, raw_value in filters.items():
        values = filter_values(raw_value)
        target = columns.get(column.upper())
        if not target or column.upper() not in allowed_filters or not values:
            continue
        clauses.append(f"TO_VARCHAR({quote_ident(target)}) in ({', '.join(['%s'] * len(values))})")
        params.extend(values)

    if search:
        search_cols = config["searchColumns"] or list(columns.values())
        actuals = [columns.get(col.upper()) for col in search_cols]
        actuals = [col for col in actuals if col]
        if actuals:
            clauses.append("(" + " or ".join(f"LOWER(TO_VARCHAR({quote_ident(col)})) like %s" for col in actuals) + ")")
            params.extend([f"%{search}%"] * len(actuals))

    where_sql = " where " + " and ".join(clauses) if clauses else ""
    paging_sql = ""
    if page_size is not None:
        offset = (page - 1) * page_size
        paging_sql = f" limit {page_size} offset {offset}"
    rows = snowflake_query(f"select * from {object_name}{where_sql}{paging_sql}", tuple(params))
    return [
        {
            "rowKey": row_key_for(row, source_config),
            "sourceName": source_config["name"],
            "snapshotDate": normalize_snapshot(run_date or first_present(row, source_config.get("snapshotColumns") or [])),
            "data": {**json_safe(row), "__sourceName": source_config["name"]},
            "syncedAt": datetime.now().isoformat(),
        }
        for row in rows
    ]


def ordered_columns(config: dict[str, Any], rows: list[dict[str, Any]]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for column in config["columns"]:
        if column not in seen:
            ordered.append(column)
            seen.add(column)
    for row in rows:
        for column in row.get("data", {}):
            if column not in seen:
                ordered.append(column)
                seen.add(column)
    return ordered


def display_columns(config: dict[str, Any], rows: list[dict[str, Any]]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    actual_by_upper: dict[str, str] = {}
    for row in rows:
        for column in row.get("data", {}):
            actual_by_upper.setdefault(str(column).upper(), column)

    def add(column: str | None) -> None:
        if not column or column in seen:
            return
        actual = actual_by_upper.get(str(column).upper(), column)
        if actual in seen:
            return
        ordered.append(actual)
        seen.add(actual)

    for column in config["columns"]:
        add(column)
    for column in config["editableColumns"]:
        add(column)
    for column in config["filterColumns"]:
        add(column)
    for source_config in config["sources"]:
        for column in source_config.get("keyColumns", []):
            add(column)
        for column in source_config.get("snapshotColumns", []):
            add(column)

    if ordered:
        return ordered
    return ordered_columns(config, rows)


def project_rows(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    column_set = set(columns)
    projected: list[dict[str, Any]] = []
    for row in rows:
        data = row.get("data", {})
        projected.append({
            **row,
            "data": {column: data.get(column) for column in columns if column in data or column in column_set},
        })
    return projected


def metrics_for(key: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    if key != "open-stock":
        return {"total": len(rows), "visibleRows": len(rows)}

    def value(row: dict[str, Any], column: str) -> str:
        return str(row.get("data", {}).get(column) or "").strip()

    def unique(column: str) -> int:
        return len({value(row, column) for row in rows if value(row, column)})

    return {
        "total": len(rows),
        "distributors": unique("DISTRIBUTOR NAME"),
        "scs": unique("SCS"),
        "outOfStock": sum(value(row, "In Stock (Y/N?)").upper() in {"N", "NO", "FALSE", "0"} for row in rows),
        "missingEta": sum(value(row, "ETA") == "" for row in rows),
        "missingPo": sum(value(row, "PO #") == "" for row in rows),
        "pendingManagement": sum(value(row, "Pending Management Comments") != "" for row in rows),
    }


def payload_cache_key(key: str, query: dict[str, list[str]]) -> str:
    normalized_query = {
        str(name): [str(value) for value in values]
        for name, values in sorted(query.items())
    }
    return json.dumps([key, normalized_query], sort_keys=True, separators=(",", ":"))


def clear_payload_cache() -> None:
    with _payload_cache_lock:
        _payload_cache.clear()


def cached_hub_payload(key: str, query: dict[str, list[str]]) -> dict[str, Any]:
    cache_key = payload_cache_key(key, query)
    now = time.monotonic()
    with _payload_cache_lock:
        cached = _payload_cache.get(cache_key)
        if cached and now - cached[0] < PAYLOAD_CACHE_TTL_SECONDS:
            return cached[1]
        if cached:
            _payload_cache.pop(cache_key, None)

    payload = get_hub_payload(key, query)

    with _payload_cache_lock:
        _payload_cache[cache_key] = (now, payload)
        while len(_payload_cache) > PAYLOAD_CACHE_LIMIT:
            _payload_cache.pop(next(iter(_payload_cache)))

    return payload


def get_hub_payload(key: str, query: dict[str, list[str]], *, screen: bool = True) -> dict[str, Any]:
    config = HUBS.get(key)
    if not config:
        raise KeyError("Hub not found.")
    errors: list[str] = []
    rows: list[dict[str, Any]] = []
    requested_source = query.get("source", [""])[0]
    sources = [
        source_config
        for source_config in config["sources"]
        if not requested_source
        or requested_source == source_config["name"]
        or requested_source == source_config.get("label")
    ]
    if not sources:
        sources = config["sources"]
    for source_config in sources:
        try:
            source_rows = build_rows_for_source(config, source_config, query)
            if key == "open-stock":
                if source_rows:
                    rows.extend(source_rows)
                    break
            else:
                rows.extend(source_rows)
        except Exception as exc:
            errors.append(f"{source_config['name']}: {exc}")
    if not rows and errors:
        raise RuntimeError(" | ".join(errors))
    columns = display_columns(config, rows) if screen else ordered_columns(config, rows)
    return {
        "hub": key,
        "label": config["label"],
        "description": config["description"],
        "columns": columns,
        "editableColumns": config["editableColumns"],
        "filterColumns": config["filterColumns"],
        "sources": [
            {
                "name": source_config["name"],
                "label": source_config.get("label", source_config["name"]),
                "objectName": source_config.get("resolvedObjectName") or source_config["objectName"],
            }
            for source_config in config["sources"]
        ],
        "rows": project_rows(rows, columns) if screen else rows,
        "total": len(rows),
        "metrics": metrics_for(key, rows),
        "sync": {"lastStatus": "live" if not errors else f"partial ({len(errors)} source errors)", "lastRunAt": datetime.now().isoformat()},
    }


def ensure_audit_tables() -> None:
    snowflake_query(f"create table if not exists {CHANGE_BATCH_TABLE} (BATCH_ID string, OPERATION string, RUN_DATE string, CHANGED_BY string, CHANGED_AT timestamp_ntz, AFFECTED_KEYS number, UNDONE_AT timestamp_ntz, UNDONE_BY string)")
    snowflake_query(f"create table if not exists {CHANGE_LOG_TABLE} (BATCH_ID string, OPERATION string, RUN_DATE string, DISTCODE_MOG_DIN string, COLUMN_NAME string, OLD_VALUE string, NEW_VALUE string, CHANGED_BY string, CHANGED_AT timestamp_ntz)")


def update_openstock_row(row_key: str, run_date: str, values: dict[str, Any], user_name: str) -> int:
    columns = source_columns({"name": "OPENSTOCKREPORT", "objectName": OPENSTOCK_TABLE})
    set_lines: list[str] = []
    params: list[Any] = []

    def has(column: str) -> bool:
        return column.upper() in columns

    def actual(column: str) -> str:
        return columns.get(column.upper(), column)

    def add_trimmed(column: str) -> None:
        if column not in values or not has(column):
            return
        set_lines.append(f"{quote_ident(actual(column))} = NULLIF(TRIM(%s), '')")
        params.append(values.get(column))

    add_trimmed("New Item?")
    if "In Stock (Y/N?)" in values and has("In Stock (Y/N?)"):
        set_lines.append(f"{quote_ident(actual('In Stock (Y/N?)'))} = CASE WHEN UPPER(TRIM(COALESCE(%s, ''))) IN ('Y','YES','TRUE','T','1') THEN 'Y' WHEN UPPER(TRIM(COALESCE(%s, ''))) IN ('N','NO','FALSE','F','0') THEN 'N' ELSE NULLIF(TRIM(%s), '') END")
        params.extend([values.get("In Stock (Y/N?)")] * 3)
    if "ETA" in values and has("ETA"):
        set_lines.append(f"{quote_ident(actual('ETA'))} = COALESCE(TO_VARCHAR(TRY_TO_DATE(NULLIF(TRIM(%s), ''), 'MM/DD/YYYY'), 'MM/DD/YYYY'), NULLIF(TRIM(%s), ''))")
        params.extend([values.get("ETA")] * 2)
    for column in ["PO #", "Current DC Comment", "Current SCS Comment", "Required DC Update", "Pending Management Comments"]:
        add_trimmed(column)
    if has("UPDATED_BY"):
        set_lines.append(f"{quote_ident(actual('UPDATED_BY'))} = %s")
        params.append(user_name or "Unknown")
    elif has("LAST_UPDATED_BY"):
        set_lines.append(f"{quote_ident(actual('LAST_UPDATED_BY'))} = %s")
        params.append(user_name or "Unknown")
    if has("UPDATED_AT"):
        set_lines.append(f"{quote_ident(actual('UPDATED_AT'))} = current_timestamp()")
    elif has("LAST_UPDATED_DATE"):
        set_lines.append(f"{quote_ident(actual('LAST_UPDATED_DATE'))} = current_timestamp()")
    if not set_lines:
        return 0
    return snowflake_execute(
        f"""
        update {OPENSTOCK_TABLE}
        set {', '.join(set_lines)}
        where {quote_ident(OPENSTOCK_KEY)} = %s
          and {insert_date_expr(quote_ident('INSERT_DATE'))} = %s
        """,
        tuple(params + [row_key, normalize_snapshot(run_date)]),
    )


def save_openstock_changes(body: dict[str, Any]) -> dict[str, Any]:
    run_date = normalize_snapshot(body.get("runDate"))
    user_name = body.get("userName") or "Unknown"
    changes = body.get("changes") if isinstance(body.get("changes"), list) else []
    if not changes:
        return {"batchId": "", "rowsAffected": 0, "loggedChanges": 0}

    config = HUBS["open-stock"]
    allowed = set(config["editableColumns"])
    ensure_audit_tables()
    batch_id = uuid.uuid4().hex.upper()
    snowflake_query(
        f"insert into {CHANGE_BATCH_TABLE} (BATCH_ID, OPERATION, RUN_DATE, CHANGED_BY, CHANGED_AT, AFFECTED_KEYS, UNDONE_AT, UNDONE_BY) select %s, 'INLINE_SAVE', %s, %s, current_timestamp(), %s, null, null",
        (batch_id, run_date, user_name, len(changes)),
    )
    rows_affected = 0
    logged = 0
    columns = source_columns({"name": "OPENSTOCKREPORT", "objectName": OPENSTOCK_TABLE})
    for change in changes:
        row_key = str(change.get("rowKey") or "").strip()
        values = {column: value for column, value in (change.get("values") or {}).items() if column in allowed}
        if not row_key or not values:
            continue
        before_cols = [columns.get(column.upper()) for column in values]
        before_cols = [column for column in before_cols if column]
        before_rows = snowflake_query(
            f"select {', '.join(quote_ident(column) for column in before_cols) if before_cols else 'null as NO_WRITABLE_COLUMNS'} from {OPENSTOCK_TABLE} where {quote_ident(OPENSTOCK_KEY)} = %s and {insert_date_expr(quote_ident('INSERT_DATE'))} = %s limit 1",
            (row_key, run_date),
        )
        before = before_rows[0] if before_rows else {}
        rows_affected += update_openstock_row(row_key, run_date, values, user_name)
        for column, value in values.items():
            old_value = first_present(before, [column])
            if str(old_value or "").strip() == str(value or "").strip():
                continue
            snowflake_query(
                f"insert into {CHANGE_LOG_TABLE} (BATCH_ID, OPERATION, RUN_DATE, DISTCODE_MOG_DIN, COLUMN_NAME, OLD_VALUE, NEW_VALUE, CHANGED_BY, CHANGED_AT) select %s, 'INLINE_SAVE', %s, %s, %s, %s, %s, %s, current_timestamp()",
                (batch_id, run_date, row_key, column, None if old_value is None else str(old_value), None if value is None else str(value), user_name),
            )
            logged += 1
    return {"batchId": batch_id, "rowsAffected": rows_affected, "loggedChanges": logged}


def recent_openstock_dates(limit: int = 25) -> list[str]:
    queries: list[str] = []
    for source_config in HUBS["open-stock"]["sources"]:
        columns = source_columns(source_config)
        snapshot = actual_column(columns, source_config.get("snapshotColumns") or [])
        if snapshot:
            queries.append(f"select distinct {insert_date_expr(quote_ident(snapshot))} as INSERT_DATE_KEY from {source_config['objectName']} where {quote_ident(snapshot)} is not null")
    if not queries:
        return []
    rows = snowflake_query(f"select INSERT_DATE_KEY from ({' union '.join(queries)}) where INSERT_DATE_KEY is not null and INSERT_DATE_KEY <> '' order by INSERT_DATE_KEY desc limit {max(1, min(limit, 100))}")
    return [str(row["INSERT_DATE_KEY"]) for row in rows if row.get("INSERT_DATE_KEY")]


def snowflake_today() -> str:
    rows = snowflake_query("select to_char(current_date(), 'YYYYMMDD') as TODAY")
    return str(rows[0]["TODAY"]) if rows else datetime.now().strftime("%Y%m%d")


def previous_openstock_date(run_date: str) -> str | None:
    selected = normalize_snapshot(run_date)
    for candidate in recent_openstock_dates(50):
        if candidate < selected:
            return candidate
    return None


def latest_openstock_undo_batch(run_date: str, user_name: str, operation: str = "INLINE_SAVE") -> str | None:
    ensure_audit_tables()
    rows = snowflake_query(
        f"""
        select BATCH_ID
        from {CHANGE_BATCH_TABLE}
        where RUN_DATE = %s
          and OPERATION = %s
          and CHANGED_BY = %s
          and UNDONE_AT is null
        order by CHANGED_AT desc
        limit 1
        """,
        (normalize_snapshot(run_date), operation, user_name or "Unknown"),
    )
    return str(rows[0]["BATCH_ID"]) if rows else None


def undo_latest_openstock_change(body: dict[str, Any]) -> dict[str, Any]:
    run_date = normalize_snapshot(body.get("runDate"))
    user_name = body.get("userName") or "Unknown"
    operation = body.get("operation") or "INLINE_SAVE"
    batch_id = latest_openstock_undo_batch(run_date, user_name, operation)
    if not batch_id:
        return {"batchId": None, "keysReverted": 0, "rowsAffected": 0, "message": "No saved change batch was found to undo."}

    change_rows = snowflake_query(
        f"""
        select DISTCODE_MOG_DIN, COLUMN_NAME, OLD_VALUE
        from {CHANGE_LOG_TABLE}
        where BATCH_ID = %s
        """,
        (batch_id,),
    )
    grouped: dict[str, dict[str, Any]] = {}
    for row in change_rows:
        key = str(row.get("DISTCODE_MOG_DIN") or "").strip()
        column = str(row.get("COLUMN_NAME") or "").strip()
        if not key or not column:
            continue
        grouped.setdefault(key, {})[column] = row.get("OLD_VALUE")

    rows_affected = 0
    for row_key, values in grouped.items():
        rows_affected += update_openstock_row(row_key, run_date, values, user_name)

    snowflake_execute(
        f"""
        update {CHANGE_BATCH_TABLE}
        set UNDONE_AT = current_timestamp(),
            UNDONE_BY = %s
        where BATCH_ID = %s
        """,
        (user_name, batch_id),
    )
    return {
        "batchId": batch_id,
        "keysReverted": len(grouped),
        "rowsAffected": rows_affected,
        "message": f"Undo complete. Keys reverted: {len(grouped):,}; rows affected: {rows_affected:,}.",
    }


def stamp_openstock_refresh_date(run_date: str) -> None:
    columns = source_columns({"name": "OPENSTOCKREPORT", "objectName": OPENSTOCK_TABLE})
    refresh_col = actual_column(columns, ["REFRESH DATE", "REFRESH_DATE", "LAST_TRANSACTION_REFRESHED_AT"])
    if not refresh_col:
        return
    snowflake_execute(
        f"""
        update {OPENSTOCK_TABLE}
        set {quote_ident(refresh_col)} = current_date()
        where {insert_date_expr(quote_ident('INSERT_DATE'))} = %s
        """,
        (normalize_snapshot(run_date),),
    )


def run_openstock_weekly_refresh(body: dict[str, Any]) -> dict[str, Any]:
    user_name = str(body.get("userName") or "").strip()
    if user_name and user_name.lower() not in SUPER_USERS:
        raise PermissionError("Weekly refresh is restricted to Open Stock super users.")
    today = snowflake_today()
    from_date = normalize_snapshot(body.get("fromDate") or (recent_openstock_dates(1)[0] if recent_openstock_dates(1) else today))
    force = bool(body.get("force"))
    already_rows = snowflake_query(
        f"""
        select count(*) as ROW_COUNT
        from {OPENSTOCK_TABLE}
        where {insert_date_expr(quote_ident('INSERT_DATE'))} = %s
        """,
        (today,),
    )
    already_run = bool(already_rows and int(already_rows[0]["ROW_COUNT"] or 0) > 0)
    if already_run and not force:
        return {
            "alreadyRunToday": True,
            "fromDate": from_date,
            "runDate": today,
            "message": "Open Stock refresh has already run today.",
        }
    snowflake_query(
        f"call {REPORT_DB}.{REPORT_SCHEMA}.OPEN_STOCK_REPORT_RUN(%s, %s)",
        (from_date, today),
    )
    stamp_openstock_refresh_date(today)
    return {
        "alreadyRunToday": False,
        "fromDate": from_date,
        "runDate": today,
        "message": f"Weekly refresh completed for {from_date} through {today}.",
    }


def persist_openstock_lookback(body: dict[str, Any]) -> dict[str, Any]:
    run_date = normalize_snapshot(body.get("runDate"))
    prev_date = normalize_snapshot(body.get("previousRunDate") or previous_openstock_date(run_date) or "")
    user_name = body.get("userName") or "Unknown"
    if not prev_date:
        return {"rowsAffected": 0, "previousRunDate": None, "message": "No previous run date is available for lookback."}

    columns = source_columns({"name": "OPENSTOCKREPORT", "objectName": OPENSTOCK_TABLE})
    key_col = actual_column(columns, [OPENSTOCK_KEY])
    insert_col = actual_column(columns, ["INSERT_DATE"])
    if not key_col or not insert_col:
        raise RuntimeError("Open Stock table is missing the key or INSERT_DATE column.")

    set_lines: list[str] = []
    change_tests: list[str] = []
    for column in OPENSTOCK_CARRY_FORWARD_COLS:
        actual = actual_column(columns, [column])
        if not actual:
            continue
        db_col = f'db.{quote_ident(actual)}'
        prev_col = f'prev.{quote_ident(actual)}'
        test = f"NULLIF(TRIM(TO_VARCHAR({db_col})), '') is null and NULLIF(TRIM(TO_VARCHAR({prev_col})), '') is not null"
        set_lines.append(f"{quote_ident(actual)} = iff({test}, TO_VARCHAR({prev_col}), {db_col})")
        change_tests.append(f"({test})")

    if actual_column(columns, ["UPDATED_BY"]):
        set_lines.append(f"{quote_ident(actual_column(columns, ['UPDATED_BY']) or 'UPDATED_BY')} = %s")
    elif actual_column(columns, ["LAST_UPDATED_BY"]):
        set_lines.append(f"{quote_ident(actual_column(columns, ['LAST_UPDATED_BY']) or 'LAST_UPDATED_BY')} = %s")
    if actual_column(columns, ["UPDATED_AT"]):
        set_lines.append(f"{quote_ident(actual_column(columns, ['UPDATED_AT']) or 'UPDATED_AT')} = current_timestamp()")
    elif actual_column(columns, ["LAST_UPDATED_DATE"]):
        set_lines.append(f"{quote_ident(actual_column(columns, ['LAST_UPDATED_DATE']) or 'LAST_UPDATED_DATE')} = current_timestamp()")

    if not change_tests:
        return {"rowsAffected": 0, "previousRunDate": prev_date, "message": "No carry-forward columns exist on the Open Stock table."}

    params: list[Any] = []
    if "%s" in ", ".join(set_lines):
        params.append(user_name)
    params.extend([run_date, prev_date])
    rows = snowflake_execute(
        f"""
        update {OPENSTOCK_TABLE} as db
        set {', '.join(set_lines)}
        from {OPENSTOCK_TABLE} as prev
        where db.{quote_ident(key_col)} = prev.{quote_ident(key_col)}
          and {insert_date_expr(f'db.{quote_ident(insert_col)}')} = %s
          and {insert_date_expr(f'prev.{quote_ident(insert_col)}')} = %s
          and ({' or '.join(change_tests)})
        """,
        tuple(params),
    )
    return {
        "rowsAffected": rows,
        "previousRunDate": prev_date,
        "message": f"Lookback persisted from {prev_date} into {run_date}. Rows affected: {rows:,}.",
    }


def csv_bytes(rows: list[dict[str, Any]], columns: list[str]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(columns)
    for row in rows:
        data = row.get("data", row)
        writer.writerow([data.get(column, "") for column in columns])
    return buffer.getvalue().encode("utf-8-sig")


def workbook_bytes(sheets: dict[str, tuple[list[dict[str, Any]], list[str]]]) -> bytes:
    if not OPENPYXL_OK:
        raise RuntimeError("Excel export requires openpyxl in the local Python runtime.")
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    used_names: set[str] = set()
    for raw_name, (rows, columns) in sheets.items():
        safe_name = re.sub(r"[\[\]:*?/\\]", "_", raw_name).strip()[:31] or "Sheet"
        base_name = safe_name
        i = 2
        while safe_name in used_names:
            suffix = f"_{i}"
            safe_name = (base_name[: 31 - len(suffix)] + suffix) if len(base_name) + len(suffix) > 31 else base_name + suffix
            i += 1
        used_names.add(safe_name)
        ws = wb.create_sheet(safe_name)
        ws.append(columns)
        for row in rows:
            data = row.get("data", row)
            ws.append([json_safe(data.get(column, "")) for column in columns])
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E9EDF2")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column in enumerate(columns, start=1):
            values = [str(column)]
            for row_idx in range(2, min(ws.max_row, 250) + 1):
                values.append(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(60, max(len(v) for v in values) + 2))
            for row_idx in range(2, min(ws.max_row, 2000) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)
        ws.auto_filter.ref = ws.dimensions
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def hub_export_bytes(key: str, query: dict[str, list[str]], export_format: str) -> tuple[bytes, str, str]:
    config = HUBS.get(key)
    if not config:
        raise KeyError("Hub not found.")
    query = {**query, "pageSize": ["all"], "page": ["1"]}
    requested_source = query.get("source", [""])[0]
    sources = [
        source_config
        for source_config in config["sources"]
        if not requested_source or requested_source in {source_config["name"], source_config.get("label", "")}
    ] or config["sources"]

    if export_format == "csv":
        payload = get_hub_payload(key, query, screen=False)
        return csv_bytes(payload["rows"], payload["columns"]), f"{key}_{datetime.now().strftime('%Y%m%d')}.csv", "text/csv; charset=utf-8"

    sheets: dict[str, tuple[list[dict[str, Any]], list[str]]] = {}
    for source_config in sources:
        source_query = {**query, "source": [source_config["name"]]}
        payload = get_hub_payload(key, source_query, screen=False)
        sheets[source_config.get("label", source_config["name"])] = (payload["rows"], payload["columns"])

    if export_format == "xlsx":
        return workbook_bytes(sheets), f"{key}_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for sheet_name, (rows, columns) in sheets.items():
            file_name = re.sub(r"[^A-Za-z0-9._-]+", "_", sheet_name).strip("_") or "export"
            zf.writestr(f"{file_name}_{datetime.now().strftime('%Y%m%d')}.csv", csv_bytes(rows, columns))
    return buffer.getvalue(), f"{key}_{datetime.now().strftime('%Y%m%d')}_csv.zip", "application/zip"


def submit_feedback(body: dict[str, Any]) -> dict[str, Any]:
    columns = source_columns({"name": "FEEDBACK_HUB", "objectName": FEEDBACK_HUB_TABLE})
    if not columns:
        raise RuntimeError("FEEDBACK_HUB table was not found or has no writable columns.")

    app_name = body.get("appName") or body.get("hub") or "Unknown"
    rating = int(body.get("rating") or 0)
    feedback_text = body.get("feedbackText") or body.get("feedback_text")
    submitted_by = body.get("submittedBy") or body.get("userName") or "Unknown"
    context_json = json.dumps(body.get("context") or {"source": "local_preview_feedback", "hub_choice": app_name})
    insert_parts: list[tuple[str, str, Any | None]] = []
    for candidate in ["EVENT_TS", "CREATED_AT", "INSERTED_AT", "UPDATED_AT", "INSERT_DATE", "CREATED_DATE"]:
        actual = actual_column(columns, [candidate])
        if actual:
            insert_parts.append((actual, "current_timestamp()", None))
            break

    values = {
        "APP_NAME": app_name,
        "RATING": rating,
        "FEEDBACK_TEXT": feedback_text,
        "SUBMITTED_BY": submitted_by,
        "SESSION_ID": str(body.get("sessionId") or uuid.uuid4()),
        "PAGE_NAME": app_name,
        "CONTEXT_JSON": context_json,
        "IS_ACTIVE": True,
    }
    for column, value in values.items():
        actual = actual_column(columns, [column])
        if not actual:
            continue
        expr = "parse_json(%s)" if column == "CONTEXT_JSON" else "%s"
        insert_parts.append((actual, expr, value))
    if not insert_parts:
        raise RuntimeError("FEEDBACK_HUB has no recognized writable columns.")
    column_sql = ", ".join(quote_ident(column) for column, _, _ in insert_parts)
    value_sql = ", ".join(expr for _, expr, _ in insert_parts)
    params = tuple(value for _, expr, value in insert_parts if "%s" in expr)
    snowflake_execute(f"insert into {FEEDBACK_HUB_TABLE} ({column_sql}) select {value_sql}", params)
    return {"ok": True, "message": "Feedback submitted."}


def resolve_hub_source(hub_key: str, source_name: str | None = None) -> dict[str, Any]:
    config = HUBS[hub_key]
    if source_name:
        for source_config in config["sources"]:
            if source_name in {source_config["name"], source_config.get("label", "")}:
                return source_config
    return config["sources"][0]


def apply_source_updates(hub_key: str, body: dict[str, Any]) -> dict[str, Any]:
    config = HUBS.get(hub_key)
    if not config:
        raise KeyError("Hub not found.")
    source_config = resolve_hub_source(hub_key, body.get("source"))
    object_name = resolve_source_object(source_config)
    columns = source_columns(source_config)
    key_col = actual_column(columns, source_config.get("keyColumns") or [])
    if not key_col:
        raise RuntimeError("This source does not expose a stable key column for direct updates.")
    allowed = {column.upper(): column for column in config.get("editableColumns") or []}
    changes = body.get("changes") if isinstance(body.get("changes"), list) else []
    user_name = body.get("userName") or "Unknown"
    rows_affected = 0
    updated = 0
    for change in changes:
        row_key = str(change.get("rowKey") or "").strip()
        raw_values = change.get("values") or {}
        values = {
            (columns.get(column.upper()) or column): value
            for column, value in raw_values.items()
            if column.upper() in allowed and columns.get(column.upper())
        }
        if not row_key or not values:
            continue
        set_lines = [f"{quote_ident(column)} = %s" for column in values]
        params = list(values.values())
        if actual_column(columns, ["UPDATED_BY"]):
            set_lines.append(f"{quote_ident(actual_column(columns, ['UPDATED_BY']) or 'UPDATED_BY')} = %s")
            params.append(user_name)
        elif actual_column(columns, ["LAST_UPDATED_BY"]):
            set_lines.append(f"{quote_ident(actual_column(columns, ['LAST_UPDATED_BY']) or 'LAST_UPDATED_BY')} = %s")
            params.append(user_name)
        if actual_column(columns, ["UPDATED_AT"]):
            set_lines.append(f"{quote_ident(actual_column(columns, ['UPDATED_AT']) or 'UPDATED_AT')} = current_timestamp()")
        elif actual_column(columns, ["LAST_UPDATED_DATE"]):
            set_lines.append(f"{quote_ident(actual_column(columns, ['LAST_UPDATED_DATE']) or 'LAST_UPDATED_DATE')} = current_timestamp()")
        params.append(row_key)
        rows_affected += snowflake_execute(
            f"update {object_name} set {', '.join(set_lines)} where TO_VARCHAR({quote_ident(key_col)}) = %s",
            tuple(params),
        )
        updated += 1
    return {"updated": updated, "rowsAffected": rows_affected, "message": f"Saved {updated:,} keyed update(s). Rows affected: {rows_affected:,}."}


def unlocked_source_pair() -> tuple[dict[str, Any], dict[str, Any]]:
    return resolve_hub_source("unlocked-accounts", "UNLOCKED_ACCOUNTS"), resolve_hub_source("unlocked-accounts", "LOCKED_INACTIVE_ACCOUNTS")


def move_unlocked_accounts(body: dict[str, Any], *, lock: bool) -> dict[str, Any]:
    unlocked_source, locked_source = unlocked_source_pair()
    source_config = unlocked_source if lock else locked_source
    target_config = locked_source if lock else unlocked_source
    source_object = resolve_source_object(source_config)
    target_object = resolve_source_object(target_config)
    source_cols = source_columns(source_config)
    target_cols = source_columns(target_config)
    id_col = actual_column(source_cols, ["ACCOUNT_RECORD_ID"])
    if not id_col:
        raise RuntimeError("ACCOUNT_RECORD_ID is required for lock/unlock moves.")
    row_keys = [str(value).strip() for value in body.get("rowKeys") or [] if str(value).strip()]
    if not row_keys:
        return {"moved": 0, "message": "No rows were selected."}
    placeholders = ", ".join(["%s"] * len(row_keys))
    target_order = list(target_cols.values())
    select_exprs: list[str] = []
    user_name = body.get("userName") or "Unknown"
    for column in target_order:
        source_actual = source_cols.get(column.upper())
        expr = f"S.{quote_ident(source_actual)}" if source_actual else "null"
        if lock and column.upper() == "LOCKED_ADDED_AT":
            expr = "current_timestamp()"
        elif lock and column.upper() == "LOCKED_ADDED_BY":
            expr = "%s"
        elif lock and column.upper() == "LOCK_REASON":
            expr = "'MANUAL_LOCK'"
        elif lock and column.upper() == "DATE_LOCKED":
            expr = "current_date()"
        elif lock and column.upper() == "DATE_UNLOCKED":
            expr = "null"
        elif not lock and column.upper() == "UNLOCKED_ADDED_AT":
            expr = "current_timestamp()"
        elif not lock and column.upper() == "UNLOCKED_ADDED_BY":
            expr = "%s"
        elif not lock and column.upper() == "DATE_UNLOCKED":
            expr = "current_date()"
        elif not lock and column.upper() == "DATE_LOCKED":
            expr = "null"
        elif column.upper() in {"LAST_UPDATED_DATE", "LAST_UPDATED_AT"}:
            expr = "current_timestamp()"
        elif column.upper() == "LAST_UPDATED_BY":
            expr = "%s"
        select_exprs.append(expr)

    params: list[Any] = []
    params.extend([user_name for expr in select_exprs if expr == "%s"])
    params.extend(row_keys)
    conn = snowflake_connect()
    try:
        with conn.cursor() as cur:
            cur.execute("begin")
            cur.execute(
                f"""
                insert into {target_object} ({', '.join(quote_ident(column) for column in target_order)})
                select {', '.join(select_exprs)}
                from {source_object} S
                where TO_VARCHAR(S.{quote_ident(id_col)}) in ({placeholders})
                """,
                tuple(params),
            )
            inserted = int(cur.rowcount or 0)
            cur.execute(
                f"delete from {source_object} where TO_VARCHAR({quote_ident(id_col)}) in ({placeholders})",
                tuple(row_keys),
            )
            cur.execute("commit")
    except Exception:
        with conn.cursor() as cur:
            cur.execute("rollback")
        raise
    action = "locked" if lock else "unlocked"
    return {"moved": inserted, "message": f"{inserted:,} account row(s) {action}."}


def bl_transfer_accounts(body: dict[str, Any]) -> dict[str, Any]:
    unlocked_source, _ = unlocked_source_pair()
    object_name = resolve_source_object(unlocked_source)
    columns = source_columns(unlocked_source)
    id_col = actual_column(columns, ["ACCOUNT_RECORD_ID"])
    dc_col = actual_column(columns, ["DC_NAME"])
    code_col = actual_column(columns, ["DISTRIBUTOR_CODE"])
    if not id_col or not dc_col or not code_col:
        raise RuntimeError("ACCOUNT_RECORD_ID, DC_NAME, and DISTRIBUTOR_CODE are required for BL transfer.")
    dest_dc = str(body.get("destinationDc") or "").strip()
    row_keys = [str(value).strip() for value in body.get("rowKeys") or [] if str(value).strip()]
    if not dest_dc or not row_keys:
        return {"rowsAffected": 0, "message": "Select rows and a destination DC first."}
    ref = snowflake_query(
        f"""
        select SUPPLY_CHAIN_NAME, SUPPLY_CHAIN_CODE
        from {REPORT_DB}.{REPORT_SCHEMA}.V_DC_MATRIX
        where SUPPLY_CHAIN_NAME = %s
        limit 1
        """,
        (dest_dc,),
    )
    if not ref:
        raise RuntimeError("Selected destination DC was not found in DC Matrix.")
    dest_code = str(ref[0].get("SUPPLY_CHAIN_CODE") or "")
    set_lines = [f"{quote_ident(dc_col)} = %s", f"{quote_ident(code_col)} = %s"]
    params: list[Any] = [dest_dc, dest_code]
    if actual_column(columns, ["LAST_UPDATED_DATE"]):
        set_lines.append(f"{quote_ident(actual_column(columns, ['LAST_UPDATED_DATE']) or 'LAST_UPDATED_DATE')} = current_timestamp()")
    if actual_column(columns, ["LAST_UPDATED_BY"]):
        set_lines.append(f"{quote_ident(actual_column(columns, ['LAST_UPDATED_BY']) or 'LAST_UPDATED_BY')} = %s")
        params.append(body.get("userName") or "Unknown")
    placeholders = ", ".join(["%s"] * len(row_keys))
    params.extend(row_keys)
    rows = snowflake_execute(
        f"update {object_name} set {', '.join(set_lines)} where TO_VARCHAR({quote_ident(id_col)}) in ({placeholders})",
        tuple(params),
    )
    return {"rowsAffected": rows, "destinationDc": dest_dc, "destinationCode": dest_code, "message": f"BL transfer updated {rows:,} row(s) to {dest_dc}."}


def dc_matrix_available_codes() -> dict[str, Any]:
    rows = snowflake_query(
        f"""
        select SUB.DISTRIBUTORCODE
        from (
            select distinct DH.DISTRIBUTORCODE
            from {DISTRIBUTOR_HIERARCHY_TABLE} DH
            where DH.DISTRIBUTORKEY > 0
            except
            select distinct DC.SUPPLY_CHAIN_CODE
            from {REPORT_DB}.{REPORT_SCHEMA}.V_DC_MATRIX DC
        ) SUB
        order by SUB.DISTRIBUTORCODE asc
        limit 5000
        """
    )
    return {"codes": [str(row["DISTRIBUTORCODE"]) for row in rows if row.get("DISTRIBUTORCODE")]}


def run_hub_action(hub_key: str, body: dict[str, Any]) -> dict[str, Any]:
    action = str(body.get("action") or "").strip()
    payload = body.get("payload") if isinstance(body.get("payload"), dict) else body
    if hub_key == "open-stock" and action == "weekly-refresh":
        return run_openstock_weekly_refresh(payload)
    if hub_key == "open-stock" and action == "persist-lookback":
        return persist_openstock_lookback(payload)
    if hub_key == "dc-matrix" and action == "available-codes":
        return dc_matrix_available_codes()
    if hub_key == "unlocked-accounts" and action == "lock-filtered":
        return move_unlocked_accounts(payload, lock=True)
    if hub_key == "unlocked-accounts" and action == "unlock-filtered":
        return move_unlocked_accounts(payload, lock=False)
    if hub_key == "unlocked-accounts" and action == "bl-transfer":
        return bl_transfer_accounts(payload)
    if action in {"save-source", "save-matrix", "save-manual"}:
        return apply_source_updates(hub_key, payload)
    raise KeyError(f"Unsupported action: {hub_key}/{action}")


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    return value


class Handler(SimpleHTTPRequestHandler):
    def translate_path(self, path: str) -> str:
        parsed = urlparse(path)
        cleaned = unquote(parsed.path)
        if cleaned == "/":
            cleaned = "/preview.html"
        direct = (ROOT / cleaned.lstrip("/")).resolve()
        public = (ROOT / "public" / cleaned.lstrip("/")).resolve()
        if str(direct).startswith(str(ROOT)) and direct.exists():
            return str(direct)
        if str(public).startswith(str(ROOT)) and public.exists():
            return str(public)
        return str(ROOT / "preview.html")

    def send_json(self, status: int, body: Any) -> None:
        data = json.dumps(json_safe(body)).encode("utf-8")
        use_gzip = "gzip" in (self.headers.get("accept-encoding") or "").lower()
        if use_gzip:
            data = gzip.compress(data, compresslevel=5)
        self.send_response(status)
        self.send_header("content-type", "application/json; charset=utf-8")
        self.send_header("cache-control", "no-store")
        self.send_header("vary", "accept-encoding")
        if use_gzip:
            self.send_header("content-encoding", "gzip")
        self.send_header("content-length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_binary(self, status: int, data: bytes, mime: str, file_name: str) -> None:
        self.send_response(status)
        self.send_header("content-type", mime)
        self.send_header("cache-control", "no-store")
        self.send_header("content-disposition", f'attachment; filename="{file_name}"')
        self.send_header("content-length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("content-length", "0") or "0")
        if not length:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def ensure_authorized(self, path: str) -> bool:
        if not path.startswith("/api/") or is_public_api_path(path):
            return True
        if is_authorized(self.headers):
            return True
        self.send_json(401, {"error": "Authentication is required."})
        return False

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        try:
            if parsed.path.lower() == "/api/auth/config":
                self.send_json(200, public_auth_config(self.headers))
                return
            if parsed.path.lower() == "/api/health":
                self.send_json(200, {"ok": True, "service": "local-preview"})
                return
            if not self.ensure_authorized(parsed.path):
                return
            hub_match = re.fullmatch(r"/api/hubs/([^/]+)", parsed.path)
            if hub_match:
                self.send_json(200, cached_hub_payload(hub_match.group(1), query))
                return
            if parsed.path == "/api/open-stock/dates":
                dates = recent_openstock_dates(int(query.get("limit", ["25"])[0] or "25"))
                selected = query.get("runDate", [dates[0] if dates else snowflake_today()])[0]
                today = snowflake_today()
                self.send_json(
                    200,
                    {
                        "dates": dates,
                        "today": today,
                        "previousDate": previous_openstock_date(selected),
                        "selectedDate": normalize_snapshot(selected),
                    },
                )
                return
            export_match = re.fullmatch(r"/api/export/([^/]+)", parsed.path)
            if export_match:
                data, file_name, mime = hub_export_bytes(export_match.group(1), query, query.get("format", ["csv"])[0])
                self.send_binary(200, data, mime, file_name)
                return
            super().do_GET()
        except KeyError as exc:
            self.send_json(404, {"error": str(exc)})
        except PermissionError as exc:
            self.send_json(403, {"error": str(exc)})
        except Exception as exc:
            self.send_json(503, {"error": str(exc)})

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path.lower() == "/api/auth/exchange-code":
                body = self.read_json()
                self.send_json(200, exchange_auth_code(str(body.get("code") or ""), str(body.get("redirectUri") or "") or None))
                return
            if not self.ensure_authorized(parsed.path):
                return
            if parsed.path == "/api/open-stock/changes":
                result = save_openstock_changes(self.read_json())
                clear_payload_cache()
                self.send_json(200, result)
                return
            if parsed.path == "/api/open-stock/undo":
                result = undo_latest_openstock_change(self.read_json())
                clear_payload_cache()
                self.send_json(200, result)
                return
            action_match = re.fullmatch(r"/api/hub-actions/([^/]+)", parsed.path)
            if action_match:
                result = run_hub_action(action_match.group(1), self.read_json())
                clear_payload_cache()
                self.send_json(200, result)
                return
            if re.fullmatch(r"/api/sync/[^/]+", parsed.path):
                self.send_json(200, {"label": "Live Snowflake", "rowsLoaded": 0, "message": "Live query mode is active; refresh uses the current source rows."})
                return
            if parsed.path == "/api/feedback":
                self.send_json(200, submit_feedback(self.read_json()))
                return
            self.send_json(404, {"error": "Endpoint not found."})
        except KeyError as exc:
            self.send_json(404, {"error": str(exc)})
        except PermissionError as exc:
            self.send_json(403, {"error": str(exc)})
        except Exception as exc:
            self.send_json(503, {"error": str(exc)})


class LegacyRedirectHandler(SimpleHTTPRequestHandler):
    def _redirect(self) -> None:
        target = f"http://localhost:{PORT}{self.path}"
        self.send_response(308)
        self.send_header("location", target)
        self.send_header("cache-control", "no-store")
        self.end_headers()

    def do_GET(self) -> None:
        self._redirect()

    def do_POST(self) -> None:
        self._redirect()

    def log_message(self, format: str, *args: Any) -> None:
        return


def start_legacy_redirect() -> None:
    if not LEGACY_REDIRECT_PORT or LEGACY_REDIRECT_PORT == PORT:
        return
    try:
        server = ThreadingHTTPServer(("127.0.0.1", LEGACY_REDIRECT_PORT), LegacyRedirectHandler)
    except OSError as exc:
        print(f"Legacy redirect unavailable on 127.0.0.1:{LEGACY_REDIRECT_PORT}: {exc}")
        return
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    print(f"Legacy redirect: http://127.0.0.1:{LEGACY_REDIRECT_PORT} -> http://localhost:{PORT}")


if __name__ == "__main__":
    start_legacy_redirect()
    print(f"Compliance Lab Python preview: http://localhost:{PORT}")
    print(f"Snowflake account: {os.environ.get('SNOWFLAKE_ACCOUNT', '<missing>')}")
    print(f"Snowflake authenticator: {os.environ.get('SNOWFLAKE_AUTHENTICATOR', 'externalbrowser')}")
    ThreadingHTTPServer(("127.0.0.1", PORT), Handler).serve_forever()
